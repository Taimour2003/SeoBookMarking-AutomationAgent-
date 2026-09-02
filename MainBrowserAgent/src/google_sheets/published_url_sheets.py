import pandas as pd
import os
from openpyxl import Workbook, load_workbook


class PublishedUrlSheets:
    def __init__(self, sheet_name="Published_URLS.xlsx"):
        self.sheet_name = sheet_name
        self.file_setup()

    def check_file_presence(self) -> bool:
        return os.path.exists(self.sheet_name)
    
    def create_sheet(self):
        if not self.check_file_presence():
            wb = Workbook()
            ws = wb.active
            ws.append(["Published_URL"])
            wb.save(self.sheet_name)
            print(f"Sheet '{self.sheet_name}' has been created.")
        else:
            print(f"Sheet '{self.sheet_name}' already exists.")
            
            
    def file_setup(self):
        if not self.check_file_presence():
            self.create_sheet()
        else:
            print(f"Sheet '{self.sheet_name}' already exists.")

    def submitted_Url(self, published_url):
        wb = load_workbook(self.sheet_name)
        ws = wb.active

        ws.append([published_url])
        wb.save(self.sheet_name)
        print(
            f"Published URL '{published_url}' has been added to the sheet '{self.sheet_name}'."
        )

    def display_urls(self):
        if not os.path.exists(self.sheet_name):
            print(f"The sheet '{self.sheet_name}' does not exist.")
            return

        wb = load_workbook(self.sheet_name)
        ws = wb.active

        print(f"Published URLs in the sheet '{self.sheet_name}':")

        for index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            print(f"{index}: {row[0]}")

    async def download_sheet(self) -> str:
        if not os.path.exists(self.sheet_name):
            raise FileNotFoundError(f"The sheet '{self.sheet_name}' does not exist.")
        return os.path.abspath(self.sheet_name)
